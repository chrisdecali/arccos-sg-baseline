#!/usr/bin/env python3
"""
pull_arccos.py — comprehensive personal Arccos Golf extractor + strokes-gained baseline.

UNOFFICIAL: reverse-engineered Arccos backend (api.arccosgolf.com). Own account only.
No aggressive polling. May break if Arccos changes their backend.

Endpoint surface confirmed live via recon on 2026-06-06 (see ENDPOINTS below).
The official strokes-gained endpoint (/v2/sga/shots) is permission-gated (401) for
dashboard tokens, so SG is RECONSTRUCTED here from raw shot coordinates + a published
Broadie expected-strokes baseline (clearly labeled approximate — see PROVENANCE).

AUTH — never handles your password. Short-lived bearer token + user id you extract
from the Arccos web dashboard (DevTools -> Network) into ~/.arccos_creds.json:
    {"bearer_token": "eyJ...", "user_id": "..."}
Tokens live ~3h; re-extract before each run. Env vars ARCCOS_BEARER_TOKEN /
ARCCOS_USER_ID also work.

ARCHITECTURE
  --fetch : pull every endpoint into a local cache (_cache_raw/, gitignored —
            contains GPS coords + PII). Round details cached by id => idempotent.
  --build : (re)build all public outputs from the cache (GPS only with --include-gps).
  --discover : dump raw endpoint structure.
  default : --fetch then --build.

OUTPUTS (./arccos_out/, GPS excluded by default — enable with --include-gps):
  rounds_summary.csv      one row per round (scoring, GIR/FW/scramble, SG categories)
  holes.csv               one row per hole (par calibrated, proximity, SG)
  shots.csv               one row per shot (club, distances-to-pin, lie, SG; lat/lng only with --include-gps)
  clubs.csv               per-club smart distance, terrain splits, GIR%, dispersion
  handicap_history.csv    per-round Arccos category handicaps
  player_profile.json     redacted profile + bag + subscription + home course
  arccos_sg_baseline.xlsx all of the above as tabs + Baseline Summary + glossary
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone
from typing import Any, Optional

try:
    from weather import fetch_round_weather
except ImportError:  # weather.py is an optional local-only enrichment module
    def fetch_round_weather(*_args, **_kwargs) -> dict:
        return {}

# ---------------------------------------------------------------------------
# Paths / config
# ---------------------------------------------------------------------------

BASE_URL = "https://api.arccosgolf.com"
CREDS_PATH = os.path.expanduser("~/.arccos_creds.json")
HERE = os.path.dirname(os.path.abspath(__file__))
# Output dir: ./arccos_out when it exists (dev layout: script sits beside the repo
# dir), otherwise this script's own dir (so the copy committed INSIDE the repo is
# runnable after a plain `git clone` regardless of the clone dir name).
OUT_DIR = os.environ.get("GOLF_STORE") or (
    os.path.join(HERE, "arccos_out") if os.path.isdir(os.path.join(HERE, "arccos_out")) else HERE)
CACHE = os.path.join(OUT_DIR, "_cache_raw")          # gitignored (GPS + PII)
DISCOVERY_DIR = os.path.join(OUT_DIR, "_discovery")  # gitignored

_HEADER_FORMS = ["Bearer: {tok}", "Bearer {tok}"]    # nonstandard colon form first
REQUEST_DELAY_S = 0.5
# GPS columns are PRIVACY-SENSITIVE (home course location). Excluded from the
# public CSVs unless explicitly enabled (env GOLF_INCLUDE_GPS=1 or --include-gps).
INCLUDE_GPS = os.environ.get("GOLF_INCLUDE_GPS", "").lower() in ("1", "true", "yes")
GPS_COLS = {"start_lat", "start_lng", "end_lat", "end_lng", "pin_lat", "pin_lng"}


def public_cols(cols: list[str]) -> list[str]:
    return list(cols) if INCLUDE_GPS else [c for c in cols if c not in GPS_COLS]


YD_PER_M = 1.0936132983

# Confirmed-working endpoints (recon 2026-06-06). {u}=user id, {r}=round, {c}=course.
ENDPOINTS = {
    "profile":          "/users/{u}",
    "subscription":     "/users/{u}/subscription",
    "handicaps":        "/users/{u}/handicaps",            # per-round category hcps
    "handicap_latest":  "/users/{u}/handicaps/latest",
    "clubs":            "/v4/clubs/user/{u}/smart-distances",
    "achievements":     "/users/{u}/achievements",
    "rounds":           "/v2/users/{u}/rounds?limit={lim}&offSet={off}",
    "round":            "/users/{u}/rounds/{r}",
    "round_ach":        "/users/{u}/rounds/{r}/achievements",
    "course":           "/courses/{c}",
    # Discovered 2026-06-06 (the high-value set the first recon missed):
    "clubs_v6":         "/v6/users/{u}/clubs?limit=50",          # real club make/model
    "tour_summary":     "/users/{u}/tourAnalyticsSummary?isDriverHandicap=F",  # career agg
    "player_summary":   "/sga/playerProfile/{u}",                # career totals
    # REAL Arccos strokes-gained (needs BOTH goalHcp and noOfRounds, else 400):
    "dashboard":        "/sga/getDashboardAnalysis/{u}?goalHcp={g}&noOfRounds={nr}",
    "round_dash":       "/sga/getDashboardAnalysis/{u}?goalHcp={g}&roundId={r}&noOfRounds=1",
    "round_analytics":  "/analytics/{r}",                        # device/app metadata
}

# Arccos SG benchmark: 0 = scratch. (Arccos uses a proprietary negative skill scale;
# goalHcp=0 gives the most interpretable "vs scratch" strokes-gained.)
GOAL_HCP = 0

# clubType id -> name. CONFIRMED 2026-06-11 against the owner's real bag as the
# Arccos app displays it (irons 5-PW, wedges 50/54/58, Cleveland HB Soft putter):
# the iron block runs 6=5i .. 11=PW — the community arccos-export enum (5=3i,
# 12=PW) is off by one for the current API; type 12 is the PUTTER (clubs_v6
# model "HB Soft 2" confirms). Wedge ids 44/42/45 = 50/54/58 per this bag's
# distance ladder + clubId registration order; other wedge ids unverified and
# left generic. Types absent from the confirmed bag (5, 46, 49, 53, 56) are
# inferred/generic — clubs.csv make/model/distance disambiguates if it matters.
CLUBTYPE = {
    1: "Driver", 2: "3 Wood", 3: "5 Wood", 4: "Hybrid", 36: "Hybrid", 37: "Hybrid",
    35: "Driving Iron",  # utility/driving iron (e.g. TaylorMade UDI) — confirmed from clubs_v6
    5: "4 Iron",  # inferred (one below confirmed 6=5 Iron)
    6: "5 Iron", 7: "6 Iron", 8: "7 Iron", 9: "8 Iron",
    10: "9 Iron", 11: "Pitching Wedge",
    44: "50 Wedge", 42: "54 Wedge", 45: "58 Wedge",
    46: "Wedge", 49: "Wedge", 53: "Wedge", 56: "Wedge",
    12: "Putter", 14: "Putter",
}


def club_category(name: Optional[str]) -> str:
    if not name:
        return "?"
    if "Iron" in name:
        return "Iron"
    if "Wood" in name:
        return "Wood"
    if "Wedge" in name:  # "Pitching Wedge", "50 Wedge", generic "Wedge"
        return "Wedge"
    return {"Driver": "Driver", "Hybrid": "Hybrid",
            "Putter": "Putter"}.get(name, "?")

# Par inference boundaries (yards, straight-line tee->pin); calibrated to course par.
PAR3_MAX_YD = 245
PAR4_MAX_YD = 470

# ---------------------------------------------------------------------------
# Broadie expected-strokes baseline (PGA Tour benchmark), APPROXIMATE.
# strokes-to-hole-out by lie + distance. Yards except 'green' in FEET.
# Source: Mark Broadie, "Every Shot Counts" (widely reproduced). Refined values
# can be dropped in without touching the SG logic. Linear interpolation between
# gridpoints; flat extrapolation beyond ends.
# ---------------------------------------------------------------------------
BASELINE = {
    "tee": {  # par 4/5 tee shot; no published values < 100yd
        100: 2.92, 120: 2.99, 140: 2.97, 160: 2.99, 180: 3.05, 200: 3.12, 220: 3.17,
        240: 3.25, 260: 3.45, 280: 3.65, 300: 3.71, 320: 3.79, 340: 3.86, 360: 3.92,
        380: 3.96, 400: 3.99, 420: 4.02, 440: 4.08, 460: 4.17, 480: 4.28, 500: 4.41,
        520: 4.54, 540: 4.65, 560: 4.74, 580: 4.79, 600: 4.82,
    },
    "fairway": {
        10: 2.18, 20: 2.40, 30: 2.52, 40: 2.60, 50: 2.66, 60: 2.70, 70: 2.72, 80: 2.75,
        90: 2.77, 100: 2.80, 120: 2.85, 140: 2.91, 160: 2.98, 180: 3.08, 200: 3.19,
        220: 3.32, 240: 3.45, 260: 3.58, 280: 3.69, 300: 3.78, 320: 3.84, 340: 3.88,
        360: 3.95, 380: 4.03, 400: 4.11, 440: 4.27, 480: 4.42, 520: 4.58, 560: 4.74, 600: 4.89,
    },
    "rough": {
        10: 2.34, 20: 2.59, 30: 2.70, 40: 2.78, 50: 2.87, 60: 2.91, 70: 2.93, 80: 2.96,
        90: 2.99, 100: 3.02, 120: 3.08, 140: 3.15, 160: 3.23, 180: 3.31, 200: 3.42,
        220: 3.53, 240: 3.64, 260: 3.74, 280: 3.83, 300: 3.90, 320: 3.95, 340: 4.02,
        360: 4.11, 380: 4.21, 400: 4.30, 440: 4.49, 480: 4.68, 520: 4.87, 560: 5.06, 600: 5.25,
    },
    "sand": {
        10: 2.43, 20: 2.53, 30: 2.66, 40: 2.82, 50: 2.92, 60: 3.15, 70: 3.21, 80: 3.24,
        90: 3.24, 100: 3.23, 120: 3.21, 140: 3.22, 160: 3.28, 180: 3.40, 200: 3.55,
        220: 3.70, 240: 3.84, 260: 3.93, 280: 4.00, 300: 4.04, 340: 4.26, 380: 4.55,
        420: 4.83, 460: 5.11, 500: 5.40, 560: 5.82, 600: 6.10,
    },
    "recovery": {
        10: 3.45, 20: 3.51, 40: 3.71, 60: 3.83, 80: 3.84, 100: 3.80, 140: 3.80, 180: 3.82,
        220: 3.92, 260: 4.03, 300: 4.20, 340: 4.44, 380: 4.66, 420: 4.84, 460: 5.03,
        500: 5.22, 560: 5.51, 600: 5.70,
    },
    "green": {  # PUTTING, distance in FEET (PGA Tour SG-Putting 2010 baseline)
        1: 1.001, 2: 1.009, 3: 1.053, 4: 1.147, 5: 1.256, 6: 1.357, 7: 1.443, 8: 1.515,
        9: 1.575, 10: 1.626, 11: 1.669, 12: 1.705, 13: 1.737, 14: 1.765, 15: 1.790,
        16: 1.811, 17: 1.830, 18: 1.848, 19: 1.863, 20: 1.878, 25: 1.934, 30: 1.978,
        35: 2.016, 40: 2.055, 45: 2.094, 50: 2.135, 60: 2.218, 70: 2.293, 80: 2.349,
        90: 2.379, 100: 2.382,
    },
}
BASELINE_SOURCE = ("Broadie, 'Assessing Golfer Performance on the PGA TOUR' Table 9 "
                   "(2003-2010 ShotLink) + PGA Tour SG-Putting 2010 baseline; same as "
                   "'Every Shot Counts'.")


def exp_strokes(lie: str, dist_yd: float) -> Optional[float]:
    """Interpolated expected strokes to hole out. dist in yards (green uses feet)."""
    table = BASELINE.get(lie)
    if not table or dist_yd is None:
        return None
    x = dist_yd * 3.0 if lie == "green" else dist_yd  # green table is in feet
    keys = sorted(table)
    if x <= keys[0]:
        return table[keys[0]]
    if x >= keys[-1]:
        return table[keys[-1]]
    for a, b in zip(keys, keys[1:]):
        if a <= x <= b:
            t = (x - a) / (b - a)
            return round(table[a] + t * (table[b] - table[a]), 4)
    return table[keys[-1]]


# ---------------------------------------------------------------------------
# Credentials (read-only; never printed)
# ---------------------------------------------------------------------------

AUTH_BASE = "https://authentication.arccosgolf.com"


def _strip_bearer(tok: Optional[str]) -> Optional[str]:
    if not tok:
        return tok
    tok = tok.strip()
    for prefix in ("Bearer:", "Bearer"):
        if tok.startswith(prefix):
            tok = tok[len(prefix):].strip()
    return tok


def _refresh_jwt(access_key: str, user_id: str) -> Optional[str]:
    """Mint a fresh short-lived JWT from a long-lived accessKey (what the
    dashboard does). POST /tokens {accessKey, userId} -> {token}. Enables
    unattended cron: the accessKey never expires, the JWT is refreshed each run."""
    body = json.dumps({"accessKey": access_key, "userId": user_id}).encode()
    req = urllib.request.Request(f"{AUTH_BASE}/tokens", data=body, method="POST")
    req.add_header("Content-Type", "application/json;charset=utf-8")
    req.add_header("Accept", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            d = json.loads(r.read().decode("utf-8"))
    except Exception as e:  # noqa: BLE001
        print(f"Warning: accessKey refresh failed ({type(e).__name__}: {e})", file=sys.stderr)
        return None
    return _strip_bearer(d.get("token") or d.get("accessToken") or d.get("jwt"))


def arccos_login(email: str, password: str) -> tuple[Optional[str], Optional[str]]:
    """Email + password -> (accessKey, userId) via POST /accessKeys. NO DevTools.
    The accessKey is long-lived; store it once and refresh JWTs from it after."""
    body = json.dumps({"email": email, "password": password,
                       "signedInByFacebook": "F"}).encode()
    req = urllib.request.Request(f"{AUTH_BASE}/accessKeys", data=body, method="POST")
    req.add_header("Content-Type", "application/json;charset=utf-8")
    req.add_header("Accept", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            d = json.loads(r.read().decode("utf-8"))
    except Exception as e:  # noqa: BLE001
        print(f"Error: Arccos login failed ({type(e).__name__}: {e})", file=sys.stderr)
        return None, None
    return d.get("accessKey"), (str(d.get("userId")) if d.get("userId") else None)


def _save_creds(updates: dict) -> None:
    cur = {}
    if os.path.exists(CREDS_PATH):
        try:
            with open(CREDS_PATH, encoding="utf-8") as f:
                cur = json.load(f)
        except (OSError, json.JSONDecodeError):
            cur = {}
    cur.update(updates)
    with open(CREDS_PATH, "w", encoding="utf-8") as f:
        json.dump(cur, f)
    os.chmod(CREDS_PATH, 0o600)


def interactive_login() -> None:
    """`--login`: prompt for email+password (no DevTools), fetch + store the
    accessKey (password is NOT stored). After this, runs are hands-off."""
    import getpass
    email = input("Arccos email: ").strip()
    password = getpass.getpass("Arccos password (hidden, NOT stored): ")
    ak, uid = arccos_login(email, password)
    if not ak:
        sys.exit("Login failed — check your email/password and try again.")
    _save_creds({"access_key": ak, **({"user_id": uid} if uid else {})})
    print(f"✓ Logged in. accessKey saved to {CREDS_PATH} (password not stored). "
          "You're set — pulls are now hands-off.")


def load_creds() -> tuple[str, str]:
    """Return (jwt, user_id). Priority: stored access_key (refresh) > email+password
    (login -> accessKey, no DevTools) > pasted bearer_token. accessKey/JWT auto-refresh
    keeps the cron hands-off."""
    token = _strip_bearer(os.environ.get("ARCCOS_BEARER_TOKEN"))
    access_key = os.environ.get("ARCCOS_ACCESS_KEY")
    user_id = os.environ.get("ARCCOS_USER_ID")
    email = os.environ.get("ARCCOS_EMAIL")
    password = os.environ.get("ARCCOS_PASSWORD")
    if os.path.exists(CREDS_PATH):
        try:
            with open(CREDS_PATH, encoding="utf-8") as f:
                c = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            sys.exit(f"Error: could not read {CREDS_PATH}: {e}")
        access_key = access_key or c.get("access_key") or c.get("accessKey")
        token = token or _strip_bearer(c.get("bearer_token") or c.get("token") or c.get("access_token"))
        user_id = user_id or c.get("user_id") or c.get("userId") or c.get("id")
        email = email or c.get("email")
        password = password or c.get("password")

    # No-DevTools path: email+password -> accessKey (only if we don't have one).
    if not access_key and email and password:
        access_key, uid = arccos_login(email, password)
        user_id = user_id or uid
    # Long-lived accessKey -> fresh short JWT each run.
    if access_key and user_id:
        fresh = _refresh_jwt(access_key, str(user_id))
        if fresh:
            token = fresh
    if not token or not user_id:
        sys.exit("Error: missing Arccos creds. Easiest: run `python3 pull_arccos.py "
                 "--login` (email+password, no DevTools). Or put access_key+user_id "
                 f"(or bearer_token+user_id) in {CREDS_PATH}.")
    return token, str(user_id)


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------

_working_header_form: Optional[str] = None


def _request(url: str, token: str, form: str) -> Any:
    req = urllib.request.Request(url)
    req.add_header("Authorization", form.format(tok=token))
    req.add_header("Content-Type", "application/json;charset=utf-8")
    req.add_header("Accept", "application/json")
    with urllib.request.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read().decode("utf-8"))


_RETRY_CODES = (500, 502, 503, 504)


def _with_retry(fn, attempts: int = 3, base_delay: float = 2.0):
    """Retry transient failures (5xx, timeouts, connection drops) with backoff.
    4xx and other HTTPErrors raise immediately. Duplicated in both pullers on
    purpose — no shared module, the ~ copies are symlinks (see plan Task 8)."""
    for i in range(attempts):
        try:
            return fn()
        except urllib.error.HTTPError as e:
            if e.code not in _RETRY_CODES or i == attempts - 1:
                raise
        except (urllib.error.URLError, TimeoutError, OSError):
            if i == attempts - 1:
                raise
        time.sleep(base_delay * (2 ** i))


def api_get(path: str, token: str, soft: bool = False) -> Any:
    """GET -> JSON. soft=True returns None on any error instead of exiting."""
    global _working_header_form
    url = path if path.startswith("http") else f"{BASE_URL}{path}"
    forms = [_working_header_form] if _working_header_form else list(_HEADER_FORMS)
    last = None
    for form in forms:
        try:
            time.sleep(REQUEST_DELAY_S)
            data = _with_retry(lambda: _request(url, token, form))
            _working_header_form = form
            return data
        except urllib.error.HTTPError as e:
            last = e
            if e.code == 401:
                continue
            if soft:
                return None
            if e.code == 429:
                sys.exit("Error: 429 Too Many Requests — back off.")
            sys.exit(f"Error: HTTP {e.code} for {url}")
        except Exception as e:  # noqa: BLE001
            last = e
            if soft:
                return None
            sys.exit(f"Error: {type(e).__name__} for {url}: {e}")
    if soft:
        return None
    sys.exit("Error: 401 for every auth form — token likely expired (~3h life). "
             f"Re-extract from the dashboard. (last: {last})")


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def haversine_yd(lat1, lon1, lat2, lon2) -> Optional[float]:
    if None in (lat1, lon1, lat2, lon2):
        return None
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return round(2 * r * math.asin(math.sqrt(a)) * YD_PER_M, 1)


# ---------------------------------------------------------------------------
# FETCH -> cache
# ---------------------------------------------------------------------------

def _save(path: str, obj: Any) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2)
    os.replace(tmp, path)


def _load(path: str) -> Any:
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def fetch_all(token: str, uid: str, n: int, after: Optional[str]) -> None:
    os.makedirs(CACHE, exist_ok=True)
    rounds_dir = os.path.join(CACHE, "rounds")
    os.makedirs(rounds_dir, exist_ok=True)

    print("Fetching account-level endpoints...")
    for key in ("profile", "subscription", "handicaps", "handicap_latest",
                "clubs", "clubs_v6", "achievements", "tour_summary", "player_summary"):
        d = api_get(ENDPOINTS[key].format(u=uid), token, soft=True)
        _save(os.path.join(CACHE, f"{key}.json"), d)
        print(f"  {key}: {'ok' if d is not None else 'MISSING'}")

    # Round summaries (paginated, newest first).
    print("Fetching round list...")
    summaries, off, page = [], 0, 50
    total = None
    while len(summaries) < n:
        url = ENDPOINTS["rounds"].format(u=uid, lim=min(page, n - len(summaries)), off=off)
        batch = api_get(url, token)
        rs = batch.get("rounds") or []
        total = batch.get("totalCount")
        if not rs:
            break
        stop = False
        for r in rs:
            if r.get("isDeleted") == "T":
                continue
            if after and (r.get("startTime") or "")[:10] < after:
                stop = True
                break
            summaries.append(r)
        off += len(rs)
        if stop or off >= (total or 0):
            break
    _save(os.path.join(CACHE, "rounds_list.json"), summaries)
    print(f"  totalCount={total}, in scope={len(summaries)}")

    # REAL Arccos strokes-gained, rolling window over the pulled rounds.
    nr = max(1, len(summaries))
    dash = api_get(ENDPOINTS["dashboard"].format(u=uid, g=GOAL_HCP, nr=nr), token, soft=True)
    if dash is None:  # some accounts reject goalHcp=0; fall back to -10
        dash = api_get(ENDPOINTS["dashboard"].format(u=uid, g=-10, nr=nr), token, soft=True)
    _save(os.path.join(CACHE, "dashboard.json"), dash)
    print(f"  dashboard SG: {'ok' if dash is not None else 'MISSING'}")

    # Course + per-round detail (cache detail by id => idempotent).
    courses_dir = os.path.join(CACHE, "courses")
    os.makedirs(courses_dir, exist_ok=True)
    seen_courses = set()
    for i, r in enumerate(summaries, 1):
        rid, cid = r.get("roundId"), r.get("courseId")
        rp = os.path.join(rounds_dir, f"{rid}.json")
        if os.path.exists(rp):
            print(f"  [{i}/{len(summaries)}] round {rid} cached")
        else:
            print(f"  [{i}/{len(summaries)}] round {rid} fetching detail + achievements")
            _save(rp, api_get(ENDPOINTS["round"].format(u=uid, r=rid), token, soft=True))
            _save(os.path.join(rounds_dir, f"{rid}_ach.json"),
                  api_get(ENDPOINTS["round_ach"].format(u=uid, r=rid), token, soft=True))
        # Per-round device metadata + REAL Arccos SG (fetch if missing => idempotent).
        adp = os.path.join(rounds_dir, f"{rid}_analytics.json")
        ddp = os.path.join(rounds_dir, f"{rid}_dash.json")
        if not os.path.exists(adp):
            _save(adp, api_get(ENDPOINTS["round_analytics"].format(r=rid), token, soft=True))
        if not os.path.exists(ddp):
            _save(ddp, api_get(ENDPOINTS["round_dash"].format(u=uid, g=GOAL_HCP, r=rid), token, soft=True))
        if cid and cid not in seen_courses:
            seen_courses.add(cid)
            cp = os.path.join(courses_dir, f"{cid}.json")
            if not os.path.exists(cp):
                _save(cp, api_get(ENDPOINTS["course"].format(c=cid), token, soft=True))
    print(f"Cache -> {CACHE}")


# ---------------------------------------------------------------------------
# Helpers for BUILD
# ---------------------------------------------------------------------------

def _tf(v) -> Optional[int]:
    return 1 if v == "T" else 0 if v == "F" else None


def clubs_v6_meta(clubs_v6: Any) -> dict[int, dict]:
    """clubId -> {type, make, model, label} from /v6 clubs (authoritative names)."""
    out: dict[int, dict] = {}
    clubs = (clubs_v6 or {}).get("clubs") if isinstance(clubs_v6, dict) else None
    if not isinstance(clubs, dict):
        return out
    for c in (clubs.get("paired") or []) + (clubs.get("unpaired") or []):
        ct = c.get("clubType")
        out[c.get("clubId")] = {
            "type": ct, "make": c.get("clubMake"), "model": c.get("clubModel"),
            "label": CLUBTYPE.get(ct, f"Club {ct}"),
        }
    return out


def build_clubid_map(meta: dict[int, dict], rounds: list[dict]) -> dict[int, str]:
    """clubId -> label. Authoritative from /v6; shot-derived fallback for any gap."""
    from collections import Counter
    m = {cid: d["label"] for cid, d in meta.items()}
    by_id: dict[int, Counter] = {}
    for rd in rounds:
        for h in (rd.get("holes") or []):
            if not h:
                continue
            for s in (h.get("shots") or []):
                if not s:
                    continue
                cid, ct = s.get("clubId"), s.get("clubType")
                if cid is None or ct is None:
                    continue
                by_id.setdefault(cid, Counter())[ct] += 1
    for cid, cnt in by_id.items():
        m.setdefault(cid, CLUBTYPE.get(cnt.most_common(1)[0][0], f"club{cid}"))
    return m


def calibrate_pars(holes_geo: list[tuple[Optional[float], Optional[int]]],
                   course_par: Optional[int]) -> list[Optional[int]]:
    """Nudge inferred pars so they sum to the true course par (handles doglegs
    reading short). Adjusts the holes closest to a class boundary first."""
    pars = [p for _, p in holes_geo]
    known = [(i, yd, p) for i, (yd, p) in enumerate(holes_geo) if p is not None]
    if not course_par or not known:
        return pars
    diff = course_par - sum(p for _, _, p in known)
    if diff == 0:
        return pars
    # Ambiguity = closeness of yardage to the nearest par boundary.
    def ambiguity(yd):
        return min(abs(yd - PAR3_MAX_YD), abs(yd - PAR4_MAX_YD)) if yd else 1e9
    order = sorted(known, key=lambda t: ambiguity(t[1]))
    step = 1 if diff > 0 else -1
    for _ in range(abs(diff)):
        for idx, _yd, _p in order:
            np_ = (pars[idx] or 0) + step
            if 3 <= np_ <= 5:
                pars[idx] = np_
                break
    return pars


def shot_lie(idx: int, n_shots: int, n_putts: int, hole: dict,
             shot: dict, par: Optional[int], start_dist_yd: Optional[float]) -> str:
    """Approximate the lie a shot was played from. APPROXIMATE — the API does not
    label per-shot lie; only tee/green are reliable."""
    if idx == 0:
        return "tee"
    if idx >= n_shots - n_putts and n_putts > 0:
        return "green"
    if shot.get("isSandUser") == "T" or shot.get("isSandUser") is True:
        return "sand"
    if shot.get("shouldConsiderPuttAsChip") == "T":
        return "fairway"  # fringe/just off green
    if idx == 1:  # second shot: fairway iff tee shot found fairway
        return "fairway" if _tf(hole.get("isFairWay")) == 1 else "rough"
    return "rough"  # later shots: conservative


# ---------------------------------------------------------------------------
# BUILD outputs
# ---------------------------------------------------------------------------

ROUND_COLS = [
    "round_id", "source", "date", "course", "tee_name", "tee_yards", "slope", "rating",
    "holes", "score", "par", "score_to_par", "pace_of_play",
    "putts", "one_putts", "three_putts", "putts_per_gir",
    "gir_hits", "gir_pct", "fairway_hits", "fairway_chances", "fairway_pct",
    "scramble_chances", "scramble_saves", "scramble_pct",
    "sand_chances_native", "sand_saves_native", "penalties",
    "avg_drive_yd", "longest_drive_yd", "avg_approach_proximity_yd",
    # REAL Arccos strokes-gained (vs scratch, goalHcp=0):
    "sg_total_arccos", "sg_off_tee_arccos", "sg_approach_arccos",
    "sg_short_arccos", "sg_putting_arccos",
    # Independent Broadie reconstruction (cross-check; may differ on putting):
    "sg_total_broadie", "sg_off_tee_broadie", "sg_approach_broadie",
    "sg_short_broadie", "sg_putting_broadie",
    "user_hcp", "drive_hcp", "approach_hcp", "chip_hcp", "sand_hcp", "putt_hcp",
    # Weather columns (external: Open-Meteo historical reanalysis by course lat/lng
    # + round mid-time UTC; non-GPS — always published regardless of GOLF_INCLUDE_GPS).
    "temp_f", "wind_mph", "wind_dir_deg", "wind_dir", "weather",
]
HOLE_COLS = [
    "round_id", "source", "date", "course", "hole_id", "par", "par_source", "shots", "net_score",
    "score_to_par", "putts", "penalties", "gir", "fairway_hit", "fw_miss_left",
    "fw_miss_right", "updown_chance_native", "updown_native", "sand_chance_native",
    "sand_save_native", "hole_len_yd", "drive_yd", "approach_proximity_yd",
    "pin_lat", "pin_lng", "scramble_chance", "scramble_save", "sg_hole_broadie",
]
SHOT_COLS = [
    "round_id", "source", "date", "hole_id", "shot_num", "club", "club_category",
    "shot_distance_yd", "start_dist_to_pin_yd", "end_dist_to_pin_yd",
    "start_lat", "start_lng", "end_lat", "end_lng",
    # Elevation in metres (terrain altitude, NOT a locating coordinate — always published
    # regardless of GOLF_INCLUDE_GPS). Useful for adjusting expected carry on hilly courses.
    "start_alt", "end_alt",
    # isHalfSwing flag from Arccos (bool -> 1/0). Helps isolate full-swing club distances
    # from partial swings; useful since there is no launch monitor data in this dataset.
    "is_half_swing",
    "lie_approx", "is_tee", "is_putt", "penalties", "category_approx", "sg_shot_approx",
]
CLUB_COLS = [
    "club", "club_category", "club_make", "club_model",
    "smart_distance_yd", "normalized_yd", "tee_yd", "fairway_yd", "rough_yd", "sand_yd",
    "longest_yd", "range_low_yd", "range_high_yd", "dispersion_yd", "gir_pct", "usage_count",
]
HCP_COLS = ["round_id", "user_hcp", "drive_hcp", "approach_hcp", "chip_hcp",
            "sand_hcp", "putt_hcp"]


def build_round(summary, detail, tee, hcp, clubid_map, rdash, pulled_at,
                course_lat=None, course_lng=None):
    holes = [h for h in (detail.get("holes") or []) if h and h.get("shouldIgnore") != "T"]
    date = (summary.get("startTime") or "")[:10]
    course = summary.get("courseName") or detail.get("courseName") or "?"
    rid = summary.get("roundId")
    course_par = summary.get("par") or detail.get("par")

    # REAL per-hole par from the dashboard's holeScores (authoritative); fall back
    # to GPS-inferred + calibrated only when the dashboard didn't return it.
    overall = (rdash or {}).get("overall") or {}
    real_par = {hs.get("holeId"): hs.get("par")
                for hs in (overall.get("holeScores") or []) if hs.get("par")}
    geo = []
    for h in holes:
        shots = [s for s in (h.get("shots") or []) if s and s.get("shouldIgnore") != "T"]
        pin = (h.get("pinLat"), h.get("pinLong"))
        yd = haversine_yd(shots[0].get("startLat"), shots[0].get("startLong"), *pin) \
            if shots and None not in pin else None
        par = (3 if (yd or 0) <= PAR3_MAX_YD else 4 if (yd or 0) <= PAR4_MAX_YD else 5) if yd else None
        geo.append((yd, par))
    inferred = calibrate_pars(geo, course_par)

    # Explicit accumulators (clearer + type-clean).
    putts_t = ones = threes = gir_n = fw_n = fw_c = scr_c = scr_s = 0
    sand_c = sand_s = pen_t = gir_putts = gir_holes = 0
    drives: list[float] = []
    prox: list[float] = []
    sg_t = sg_tee = sg_app = sg_short = sg_putt = 0.0
    sg_have = False

    hole_rows, shot_rows = [], []
    for hi, h in enumerate(holes):
        hid = h.get("holeId")
        par = real_par.get(hid) or inferred[hi]
        par_source = "arccos" if real_par.get(hid) else "inferred"
        yd = geo[hi][0]
        shots = [s for s in (h.get("shots") or []) if s and s.get("shouldIgnore") != "T"]
        n = len(shots)
        putts = h.get("putts") or 0
        gir = _tf(h.get("isGir"))
        fw = _tf(h.get("isFairWay"))
        pin = (h.get("pinLat"), h.get("pinLong"))
        hole_pen = sum(s.get("noOfPenalties") or 0 for s in shots)

        putts_t += putts
        pen_t += hole_pen
        ones += 1 if putts == 1 else 0
        threes += 1 if putts >= 3 else 0
        if gir == 1:
            gir_n += 1
            gir_putts += putts
            gir_holes += 1
        if par is not None and par >= 4:
            fw_c += 1
            fw_n += 1 if fw == 1 else 0

        drive_yd = None
        if n and par and par >= 4:
            s0 = shots[0]
            drive_yd = haversine_yd(s0.get("startLat"), s0.get("startLong"),
                                    s0.get("endLat"), s0.get("endLong"))
            if drive_yd:
                drives.append(drive_yd)

        approach_prox = None
        hole_sg = 0.0
        for si, s in enumerate(shots):
            sd = haversine_yd(s.get("startLat"), s.get("startLong"), *pin) if None not in pin else None
            ed = haversine_yd(s.get("endLat"), s.get("endLong"), *pin) if None not in pin else None
            lie = shot_lie(si, n, putts, h, s, par, sd)
            holed = (si == n - 1)
            pen = s.get("noOfPenalties") or 0
            sg_shot = None
            start_e = exp_strokes(lie, sd) if sd is not None else None
            if start_e is not None:
                end_lie = "green" if (si + 1 >= n - putts and putts > 0) else "fairway"
                end_e = 0.0 if holed else (exp_strokes(end_lie, ed) if ed is not None else None)
                if end_e is not None:
                    sg_shot = round(start_e - end_e - 1 - pen, 3)
            if lie == "tee" and par and par >= 4:
                cat = "off_tee"
            elif lie == "green":
                cat = "putting"
            elif sd is not None and sd <= 30:
                cat = "short_game"
            else:
                cat = "approach"
            if sg_shot is not None:
                sg_have = True
                hole_sg += sg_shot
                sg_t += sg_shot
                if cat == "off_tee":
                    sg_tee += sg_shot
                elif cat == "approach":
                    sg_app += sg_shot
                elif cat == "short_game":
                    sg_short += sg_shot
                else:
                    sg_putt += sg_shot
            if cat in ("approach", "short_game") and si == n - putts - 1 and ed is not None:
                approach_prox = ed
                prox.append(ed)
            cname = CLUBTYPE.get(s.get("clubType")) or clubid_map.get(s.get("clubId")) or "?"
            shot_rows.append({
                "round_id": rid, "date": date, "hole_id": hid, "shot_num": si + 1,
                "club": cname, "club_category": club_category(cname),
                "shot_distance_yd": haversine_yd(s.get("startLat"), s.get("startLong"),
                                                 s.get("endLat"), s.get("endLong")),
                "start_dist_to_pin_yd": sd, "end_dist_to_pin_yd": ed,
                "start_lat": s.get("startLat"), "start_lng": s.get("startLong"),
                "end_lat": s.get("endLat"), "end_lng": s.get("endLong"),
                "start_alt": s.get("startAltitude"), "end_alt": s.get("endAltitude"),
                "is_half_swing": 1 if s.get("isHalfSwing") else 0,
                "lie_approx": lie, "is_tee": 1 if si == 0 else 0,
                "is_putt": 1 if lie == "green" else 0, "penalties": pen,
                "category_approx": cat, "sg_shot_approx": sg_shot,
            })

        scr_chance = scr_save = None
        if par is not None and gir is not None:
            scr_chance = 1 if gir == 0 else 0
            scr_save = 1 if (gir == 0 and n <= par) else 0
            scr_c += scr_chance
            scr_s += scr_save
        sand_c += _tf(h.get("isSandSaveChance")) or 0
        sand_s += _tf(h.get("isSandSave")) or 0

        hole_rows.append({
            "round_id": rid, "date": date, "course": course, "hole_id": hid,
            "par": par, "par_source": par_source, "shots": n,
            "net_score": (n - par) if par is not None else None,
            "score_to_par": (n - par) if par is not None else None,
            "putts": putts, "penalties": hole_pen, "gir": gir, "fairway_hit": fw,
            "fw_miss_left": _tf(h.get("isFairWayLeft")), "fw_miss_right": _tf(h.get("isFairWayRight")),
            "updown_chance_native": _tf(h.get("isUpDownChance")), "updown_native": _tf(h.get("isUpDown")),
            "sand_chance_native": _tf(h.get("isSandSaveChance")), "sand_save_native": _tf(h.get("isSandSave")),
            "hole_len_yd": yd, "drive_yd": drive_yd, "approach_proximity_yd": approach_prox,
            "pin_lat": pin[0], "pin_lng": pin[1],
            "scramble_chance": scr_chance, "scramble_save": scr_save,
            "sg_hole_broadie": round(hole_sg, 3) if sg_have else None,
        })

    nh = len(holes)
    score = summary.get("scoreOverride") or summary.get("noOfShots")
    sec = overall.get("overallSection") or {}

    # Weather enrichment (external: Open-Meteo historical reanalysis — not Arccos).
    # Uses course lat/lng (fetched during build) + round mid-time UTC hour.
    # Always graceful: returns {} on any failure so the pull is never broken.
    wx: dict = {}
    if course_lat is not None and course_lng is not None and date:
        try:
            start_str = summary.get("startTime") or ""
            end_str = summary.get("endTime") or ""
            _fmt = "%Y-%m-%dT%H:%M:%S.%fZ"
            def _parse_utc(s: str) -> Optional[datetime]:
                for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ"):
                    try:
                        return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
                    except ValueError:
                        continue
                return None
            t_start = _parse_utc(start_str)
            t_end = _parse_utc(end_str)
            if t_start and t_end:
                mid_ts = t_start.timestamp() + (t_end.timestamp() - t_start.timestamp()) / 2
                mid_hour = int(datetime.fromtimestamp(mid_ts, tz=timezone.utc).hour)
            elif t_start:
                mid_hour = int(t_start.hour)
            else:
                mid_hour = 12  # noon UTC fallback
            wx = fetch_round_weather(
                course_lat, course_lng, date, mid_hour, cache_dir=OUT_DIR
            )
        except Exception:  # noqa: BLE001
            wx = {}

    rr = {
        "round_id": rid, "date": date, "course": course,
        "tee_name": tee.get("name"), "tee_yards": tee.get("distance"),
        "slope": tee.get("slope"), "rating": tee.get("rating"),
        "holes": nh, "score": score, "par": course_par,
        "score_to_par": summary.get("overUnder"), "pace_of_play": overall.get("paceOfPlay"),
        "putts": putts_t, "one_putts": ones, "three_putts": threes,
        "putts_per_gir": round(gir_putts / gir_holes, 2) if gir_holes else None,
        "gir_hits": gir_n, "gir_pct": round(100 * gir_n / nh, 1) if nh else None,
        "fairway_hits": fw_n, "fairway_chances": fw_c,
        "fairway_pct": round(100 * fw_n / fw_c, 1) if fw_c else None,
        "scramble_chances": scr_c, "scramble_saves": scr_s,
        "scramble_pct": round(100 * scr_s / scr_c, 1) if scr_c else None,
        "sand_chances_native": sand_c, "sand_saves_native": sand_s, "penalties": pen_t,
        "avg_drive_yd": round(sum(drives) / len(drives), 1) if drives else None,
        "longest_drive_yd": max(drives) if drives else None,
        "avg_approach_proximity_yd": round(sum(prox) / len(prox), 1) if prox else None,
        # REAL Arccos SG (vs scratch).
        "sg_total_arccos": sec.get("sga"), "sg_off_tee_arccos": sec.get("drivingSga"),
        "sg_approach_arccos": sec.get("approachSga"), "sg_short_arccos": sec.get("shortSga"),
        "sg_putting_arccos": sec.get("puttingSga"),
        # Broadie reconstruction (cross-check).
        "sg_total_broadie": round(sg_t, 2) if sg_have else None,
        "sg_off_tee_broadie": round(sg_tee, 2) if sg_have else None,
        "sg_approach_broadie": round(sg_app, 2) if sg_have else None,
        "sg_short_broadie": round(sg_short, 2) if sg_have else None,
        "sg_putting_broadie": round(sg_putt, 2) if sg_have else None,
        "user_hcp": hcp.get("userHcp"), "drive_hcp": hcp.get("driveHcp"),
        "approach_hcp": hcp.get("approachHcp"), "chip_hcp": hcp.get("chipHcp"),
        "sand_hcp": hcp.get("sandHcp"), "putt_hcp": hcp.get("puttHcp"),
        # Weather (Open-Meteo historical reanalysis; {} when unavailable).
        "temp_f": wx.get("temp_f"), "wind_mph": wx.get("wind_mph"),
        "wind_dir_deg": wx.get("wind_dir_deg"), "wind_dir": wx.get("wind_dir"),
        "weather": wx.get("weather"),
        "pulled_at": pulled_at,
    }
    return rr, hole_rows, shot_rows


def write_csv(path, cols, rows):
    tmp = path + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in cols})
    os.replace(tmp, path)


def build_clubs_csv(clubs_raw, clubid_map, meta):
    rows = []
    for c in (clubs_raw.get("clubs") if isinstance(clubs_raw, dict) else clubs_raw) or []:
        cid = c.get("clubId")
        name = clubid_map.get(cid, f"club{cid}")
        m = meta.get(cid, {})
        def r1(x):  # round-or-None, type-clean (evaluates source once)
            return round(x, 1) if isinstance(x, (int, float)) else None
        terr = c.get("terrain") or {}
        def td(k):
            return (terr.get(k) or {}).get("distance")
        rng = c.get("range") or {}
        lo, hi = rng.get("low"), rng.get("high")
        rows.append({
            "club": name, "club_category": club_category(name),
            "club_make": m.get("make"), "club_model": (m.get("model") or "").strip() or None,
            "smart_distance_yd": r1((c.get("smartDistance") or {}).get("distance")),
            "normalized_yd": r1((c.get("normalizedSmartDistance") or {}).get("distance")),
            "tee_yd": r1(td("tee")), "fairway_yd": r1(td("fairway")),
            "rough_yd": r1(td("rough")), "sand_yd": r1(td("sand")),
            "longest_yd": r1((c.get("longest") or {}).get("distance")),
            "range_low_yd": lo, "range_high_yd": hi,
            "dispersion_yd": r1(hi - lo) if isinstance(lo, (int, float)) and isinstance(hi, (int, float)) else None,
            "gir_pct": r1((c.get("gir") or {}).get("percentage")),
            "usage_count": (c.get("usage") or {}).get("count"),
        })
    # Longest first.
    rows.sort(key=lambda r: r["smart_distance_yd"] or 0, reverse=True)
    return rows


def redact_profile(prof, sub):
    if not isinstance(prof, dict):
        return {}
    drop = {"email", "firstName", "lastName", "dateOfBirth", "fcmToken",
            "profilePicUrl", "phone"}
    out = {k: v for k, v in prof.items() if k not in drop and not isinstance(v, (list, dict))}
    hc = prof.get("homeCourse") or {}
    out["homeCourse"] = {k: hc.get(k) for k in ("name", "city", "state", "country")}
    out["preferredBall"] = prof.get("preferredBall")
    out["bag"] = prof.get("bags")
    if isinstance(sub, dict):
        out["subscription"] = sub
    out["_note"] = "Redacted: email/name/DOB/token/coords removed. handicap/userHcp use Arccos' proprietary negative scale, not USGA index."
    return out


def _pct(num, den):
    if not den or num is None:
        return None
    return round(100 * num / den, 1)


def build_career_stats(tour, player, dashboard, pulled_at, counts) -> dict:
    """Flatten tourAnalyticsSummary + playerProfile + dashboard SG into one
    aggregate file, with computed rates. Short-game fields are first-class."""
    tour = tour or {}
    drive, app = tour.get("drive") or {}, tour.get("approach") or {}
    chip, sand = tour.get("chip") or {}, tour.get("sand") or {}
    putt, ov = tour.get("putt") or {}, tour.get("overall") or {}
    do = (dashboard or {}).get("overall") or {}
    rates = {
        "fairway_pct": _pct(drive.get("noOfFairWaysHit"), drive.get("noOfFairwayAttempts")),
        "gir_pct": _pct(app.get("noOfGreensHitInRegulation"), app.get("noOfHoles")),
        "scramble_chip_save_pct": _pct(chip.get("noOfChipSaveSuccesses"), chip.get("noOfChipSaveChances")),
        "chip_down_pct": _pct(chip.get("noOfChipDownSuccesses"), chip.get("noOfChipDownChances")),
        "chip_error_rate": _pct(chip.get("noOfErrors"), chip.get("noOfChipShots")),
        "sand_save_pct": _pct(sand.get("noOfSandSaveSuccesses"), sand.get("noOfSandSaveChances")),
        "three_putt_pct": _pct(putt.get("threePlusPutts"), putt.get("noOfHoles")),
        "one_putt_pct": _pct(putt.get("onePutts"), putt.get("noOfHoles")),
        "putts_per_round": (round(total_putts / counts["rounds"], 2)
                            if (total_putts := putt.get("totalPutts")) and counts["rounds"]
                            else None),
    }
    return {
        "pulled_at": pulled_at, "rounds": counts["rounds"],
        "holes": counts["holes"], "shots": counts["shots"],
        "player_summary": player or {},
        "strokes_gained_arccos": do.get("overallSection") or {},
        "caddie_insights": do.get("caddieInsights") or {},
        "score_analysis": do.get("scoreAnalysis") or {},
        "key_rates": rates,
        "career_by_category": {"drive": drive, "approach": app, "chip": chip,
                               "sand": sand, "putt": putt, "overall": ov},
        "_note": ("Aggregate career stats from Arccos tourAnalyticsSummary + "
                  "getDashboardAnalysis (SG vs scratch) + playerProfile. *_hcp are "
                  "Arccos' proprietary negative skill scale, not USGA."),
    }


SGA_BANDS_COLS = [
    "section", "metric", "slab", "slab_unit", "terrain", "sga",
    "shots_count", "avg_dist_to_pin", "dist_to_pin_unit", "goal", "extra",
]


def extract_sga_bands(dashboard: Optional[dict]) -> list[dict]:
    """Extract Arccos-computed SG band breakdowns from the account-level dashboard
    response into a tidy long-format list. Sections/fields are Arccos-proprietary
    (pre-computed strokes-gained vs goalHcp=0 scratch benchmark). Missing sections
    are skipped gracefully — callers with few rounds may have empty subsections.

    The resulting rows feed sga_bands.csv (non-GPS, always published).
    """
    if not isinstance(dashboard, dict):
        return []
    rows: list[dict] = []

    def row(**kwargs) -> dict:
        base: dict = {c: None for c in SGA_BANDS_COLS}
        base.update(kwargs)
        return base

    def slab_val(obj: Any) -> tuple[Optional[str], Optional[str]]:
        """Extract (value, unit) from a slab/distance object like {"value": "50-100", "unit": "Yards"}."""
        if not isinstance(obj, dict):
            return None, None
        return obj.get("value"), obj.get("unit")

    driving = dashboard.get("driving") or {}
    # driving.distanceVsAccuracy — aggregate SG for distance/accuracy/penalties
    dva = driving.get("distanceVsAccuracy") or {}
    if dva:
        rows.append(row(section="driving", metric="sg_distance", sga=dva.get("sgDistance")))
        rows.append(row(section="driving", metric="sg_accuracy", sga=dva.get("sgAccuracy")))
        rows.append(row(section="driving", metric="sg_penalties", sga=dva.get("sgPenalties")))

    # driving.drivingDistance — avg distance + goal + longest
    dd = driving.get("drivingDistance") or {}
    if dd:
        avg_obj = dd.get("averageDistance") or {}
        goal_obj = dd.get("goal") or {}
        rows.append(row(
            section="driving", metric="driving_distance",
            sga=avg_obj.get("value"),
            goal=goal_obj.get("value"),
            extra=f"longestDrive={dd.get('longestDrive')},drivingCount={dd.get('drivingCount')}",
        ))

    # driving.drivingByHoleLength[] — SG by tee-shot hole length band
    for entry in (driving.get("drivingByHoleLength") or []):
        sv, su = slab_val(entry.get("slab"))
        rows.append(row(
            section="driving", metric="driving_by_hole_length",
            slab=sv, slab_unit=su,
            sga=entry.get("sga"), shots_count=entry.get("shotsCount"),
        ))

    approach = dashboard.get("approach") or {}
    # approach.approachByPinDistance[] — SG from various distances
    for entry in (approach.get("approachByPinDistance") or []):
        sv, su = slab_val(entry.get("slab"))
        rows.append(row(
            section="approach", metric="approach_by_pin_distance",
            slab=sv, slab_unit=su,
            sga=entry.get("sga"), shots_count=entry.get("shotsCount"),
        ))

    # approach.approachByTerrain[] — SG by lie (fairway/rough/sand/tee)
    for entry in (approach.get("approachByTerrain") or []):
        rows.append(row(
            section="approach", metric="approach_by_terrain",
            terrain=entry.get("terrain"),
            sga=entry.get("sga"), shots_count=entry.get("shotsCount"),
        ))

    short = dashboard.get("short") or {}
    # short.chipByPinDistance[] — chip SG by distance to pin (key chip-leave data)
    for entry in (short.get("chipByPinDistance") or []):
        sv, su = slab_val(entry.get("slab"))
        rows.append(row(
            section="short", metric="chip_by_pin_distance",
            slab=sv, slab_unit=su,
            sga=entry.get("sga"), shots_count=entry.get("shotsCount"),
        ))

    # short.chippingAccuracy[] — avg distance to pin by slab with goal
    for entry in (short.get("chippingAccuracy") or []):
        sv, su = slab_val(entry.get("slab"))
        adp = entry.get("avgDistanceToPin") or {}
        adp_goal = entry.get("avgDistanceToPinGoal") or {}
        rows.append(row(
            section="short", metric="chipping_accuracy",
            slab=sv, slab_unit=su,
            avg_dist_to_pin=adp.get("value"), dist_to_pin_unit=adp.get("unit"),
            goal=adp_goal.get("value"),
        ))

    # short.sandByPinDistance[] — bunker SG by distance
    for entry in (short.get("sandByPinDistance") or []):
        sv, su = slab_val(entry.get("slab"))
        rows.append(row(
            section="short", metric="sand_by_pin_distance",
            slab=sv, slab_unit=su,
            sga=entry.get("sga"), shots_count=entry.get("shotsCount"),
        ))

    # short.sandAccuracy[] — avg distance to pin from sand by slab
    for entry in (short.get("sandAccuracy") or []):
        sv, su = slab_val(entry.get("slab"))
        adp = entry.get("avgDistanceToPin") or {}
        adp_goal = entry.get("avgDistanceToPinGoal") or {}
        rows.append(row(
            section="short", metric="sand_accuracy",
            slab=sv, slab_unit=su,
            avg_dist_to_pin=adp.get("value"), dist_to_pin_unit=adp.get("unit"),
            goal=adp_goal.get("value"),
        ))

    putting = dashboard.get("putting") or {}
    # putting.puttingByLength[] — SG by putt length band
    for entry in (putting.get("puttingByLength") or []):
        sv, su = slab_val(entry.get("slab"))
        rows.append(row(
            section="putting", metric="putting_by_length",
            slab=sv, slab_unit=su,
            sga=entry.get("sga"), shots_count=entry.get("shotsCount"),
        ))

    # overall.overallSection.caddieInsights — helping/hurting factors
    # May be empty with few rounds; tolerate gracefully.
    overall = dashboard.get("overall") or {}
    ov_sec = overall.get("overallSection") or {}
    caddie = ov_sec.get("caddieInsights") or {}
    for metric, key in (("caddie_helping", "helping"), ("caddie_hurting", "hurting")):
        for entry in (caddie.get(key) or []):
            dist_obj = entry.get("fromDistance") or {}
            sv, _ = slab_val(dist_obj) if isinstance(dist_obj, dict) else (None, None)
            rows.append(row(
                section="overall", metric=metric,
                slab=sv,
                terrain=entry.get("from"),
                sga=entry.get("sga"),
                extra=entry.get("label"),
            ))

    return rows


def build(pulled_at: str) -> dict:
    os.makedirs(OUT_DIR, exist_ok=True)
    summaries = _load(os.path.join(CACHE, "rounds_list.json")) or []
    hcp_hist = (_load(os.path.join(CACHE, "handicaps.json")) or {}).get("handicaps") or []
    hcp_by_round = {h.get("roundId"): h for h in hcp_hist}
    clubs_raw = _load(os.path.join(CACHE, "clubs.json")) or {}
    meta = clubs_v6_meta(_load(os.path.join(CACHE, "clubs_v6.json")))

    # Load round details + course tees + per-round dashboards.
    details, courses, rdashes = {}, {}, {}
    course_latlng: dict = {}  # courseId -> (lat, lng) for weather enrichment
    for r in summaries:
        rid = r.get("roundId")
        d = _load(os.path.join(CACHE, "rounds", f"{rid}.json"))
        if d:
            details[rid] = d
        rdashes[rid] = _load(os.path.join(CACHE, "rounds", f"{rid}_dash.json"))
        cid = r.get("courseId")
        if cid and cid not in courses:
            c = _load(os.path.join(CACHE, "courses", f"{cid}.json"))
            if c:
                courses[cid] = {str(t.get("teeId")): t for t in (c.get("courseTees") or [])}
                lat, lng = c.get("latitude"), c.get("longitude")
                if lat is not None and lng is not None:
                    course_latlng[cid] = (float(lat), float(lng))

    clubid_map = build_clubid_map(meta, list(details.values()))

    round_rows, hole_rows, shot_rows, hcp_rows = [], [], [], []
    for r in summaries:
        rid = r.get("roundId")
        detail = details.get(rid)
        if not detail:
            continue
        cid = r.get("courseId")
        tee = courses.get(cid, {}).get(str(r.get("teeId")), {})
        hcp = hcp_by_round.get(rid, {})
        c_lat, c_lng = (course_latlng.get(cid) or (None, None))
        rr, hr, sr = build_round(r, detail, tee, hcp, clubid_map, rdashes.get(rid), pulled_at,
                                 course_lat=c_lat, course_lng=c_lng)
        round_rows.append(rr)
        hole_rows.extend(hr)
        shot_rows.extend(sr)
        if hcp:
            hcp_rows.append({c: hcp.get({"round_id": "roundId", "user_hcp": "userHcp",
                            "drive_hcp": "driveHcp", "approach_hcp": "approachHcp",
                            "chip_hcp": "chipHcp", "sand_hcp": "sandHcp",
                            "putt_hcp": "puttHcp"}[c]) for c in HCP_COLS})

    round_rows.sort(key=lambda r: r.get("date") or "")
    hole_rows.sort(key=lambda r: (r.get("date") or "", r.get("round_id"), r.get("hole_id")))
    shot_rows.sort(key=lambda r: (r.get("date") or "", r.get("round_id"), r.get("hole_id"), r.get("shot_num")))

    # Provenance: tag every native Arccos row so multi-source merge (import_external_rounds.py)
    # can dedup and label. Additive column; downstream reads by name and tolerates its absence.
    for _r in round_rows:
        _r["source"] = "arccos"
    for _r in hole_rows:
        _r["source"] = "arccos"
    for _r in shot_rows:
        _r["source"] = "arccos"

    write_csv(os.path.join(OUT_DIR, "rounds_summary.csv"), ROUND_COLS, round_rows)
    write_csv(os.path.join(OUT_DIR, "holes.csv"), public_cols(HOLE_COLS), hole_rows)
    write_csv(os.path.join(OUT_DIR, "shots.csv"), public_cols(SHOT_COLS), shot_rows)
    club_rows = build_clubs_csv(clubs_raw, clubid_map, meta)
    write_csv(os.path.join(OUT_DIR, "clubs.csv"), CLUB_COLS, club_rows)
    write_csv(os.path.join(OUT_DIR, "handicap_history.csv"), HCP_COLS, hcp_rows)

    prof = redact_profile(_load(os.path.join(CACHE, "profile.json")),
                          _load(os.path.join(CACHE, "subscription.json")))
    _save(os.path.join(OUT_DIR, "player_profile.json"), prof)

    # sga_bands.csv — Arccos-computed SG band breakdowns (non-GPS, always published).
    # Source: account-level dashboard.json. Bands are Arccos' own SG calculation vs goalHcp=0.
    dashboard_raw = _load(os.path.join(CACHE, "dashboard.json"))
    band_rows = extract_sga_bands(dashboard_raw)
    write_csv(os.path.join(OUT_DIR, "sga_bands.csv"), SGA_BANDS_COLS, band_rows)

    counts = dict(rounds=len(round_rows), holes=len(hole_rows), shots=len(shot_rows),
                  clubs=len(club_rows), hcp=len(hcp_rows), sga_bands=len(band_rows))
    career = build_career_stats(_load(os.path.join(CACHE, "tour_summary.json")),
                                _load(os.path.join(CACHE, "player_summary.json")),
                                dashboard_raw, pulled_at, counts)
    _save(os.path.join(OUT_DIR, "career_stats.json"), career)

    build_xlsx(round_rows, hole_rows, shot_rows, club_rows, hcp_rows, career, pulled_at, counts)
    return counts


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def _num(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def build_xlsx(rounds, holes, shots, clubs, hcps, career, pulled_at, counts):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)
    hdr = PatternFill("solid", fgColor="DDEBF7")
    hl = PatternFill("solid", fgColor="FFF2CC")

    def tab(ws, cols, rows, title):
        ws.title = title
        ws.append(cols)
        for c in ws[1]:
            c.font, c.fill = bold, hdr
        for r in rows:
            ws.append([_num(r.get(k)) if _num(r.get(k)) is not None else (r.get(k) or "") for k in cols])
        ws.freeze_panes = "A2"
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(10, min(30, len(col) + 2))

    tab(wb.active, ROUND_COLS, rounds, "Rounds")
    tab(wb.create_sheet(), public_cols(HOLE_COLS), holes, "Holes")
    tab(wb.create_sheet(), public_cols(SHOT_COLS), shots, "Shots")
    tab(wb.create_sheet(), CLUB_COLS, clubs, "Clubs")
    tab(wb.create_sheet(), HCP_COLS, hcps, "Handicap History")

    # Career Stats tab — flattened Arccos aggregate (tourAnalyticsSummary + SG).
    cs = wb.create_sheet(title="Career Stats")
    cs.column_dimensions["A"].width = 44
    cs.column_dimensions["B"].width = 16
    cs.append(["CAREER STATS (Arccos aggregate)"])
    cs["A1"].font = Font(bold=True, size=13)
    sga = career.get("strokes_gained_arccos", {}) or {}
    crows = [("", ""), ("STROKES GAINED (vs scratch)", ""),
             ("SG total", sga.get("sga")), ("SG driving", sga.get("drivingSga")),
             ("SG approach", sga.get("approachSga")), ("SG short", sga.get("shortSga")),
             ("SG putting", sga.get("puttingSga")), ("", ""), ("KEY RATES", "")]
    crows += list((career.get("key_rates") or {}).items())
    crows += [("", ""), ("CAREER BY CATEGORY (raw)", "")]
    for cat_name, cat in (career.get("career_by_category") or {}).items():
        crows.append((f"-- {cat_name} --", ""))
        for k, v in (cat or {}).items():
            if isinstance(v, (int, float)):
                crows.append((f"{cat_name}.{k}", round(v, 2) if isinstance(v, float) else v))
    for a, b in crows:
        cs.append([a, b if b != "" else None])
        if a in ("STROKES GAINED (vs scratch)", "KEY RATES", "CAREER BY CATEGORY (raw)"):
            for c in cs[cs.max_row]:
                c.font, c.fill = bold, hdr

    ws = wb.create_sheet(title="Baseline Summary")
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 14
    ws.append(["ARCCOS STROKES-GAINED BASELINE"])
    ws["A1"].font = Font(bold=True, size=14)
    dates = sorted(r["date"] for r in rounds if r.get("date"))
    ws.append(["Rounds", counts["rounds"]])
    ws.append(["Holes / Shots", f'{counts["holes"]} / {counts["shots"]}'])
    ws.append(["Date range", f"{dates[0]} -> {dates[-1]}" if dates else "—"])
    ws.append([])
    ws.append(["Metric (avg across rounds)", "Value"])
    for c in ws[ws.max_row]:
        c.font, c.fill = bold, hdr

    short_game = {"scramble_pct", "sg_short_arccos", "chip_hcp", "sand_hcp",
                  "avg_approach_proximity_yd"}
    metrics = [
        ("score", "Avg score"), ("score_to_par", "Avg score to par"),
        ("sg_total_arccos", "SG total — Arccos (vs scratch)"),
        ("sg_off_tee_arccos", "SG off-tee — Arccos"),
        ("sg_approach_arccos", "SG approach — Arccos"),
        ("sg_short_arccos", "SG short game — Arccos — PRIORITY"),
        ("sg_putting_arccos", "SG putting — Arccos"),
        ("sg_total_broadie", "SG total — Broadie reconstruction (cross-check)"),
        ("gir_pct", "GIR %"), ("fairway_pct", "Fairway %"),
        ("scramble_pct", "Scrambling % (derived) — PRIORITY"),
        ("avg_approach_proximity_yd", "Proximity, last shot to pin (yd) — PRIORITY"),
        ("putts", "Putts / round"), ("one_putts", "1-putts / round"),
        ("three_putts", "3-putts / round"), ("putts_per_gir", "Putts per GIR"),
        ("avg_drive_yd", "Avg drive (yd)"), ("longest_drive_yd", "Longest drive (yd)"),
        ("penalties", "Penalties / round"),
        ("chip_hcp", "chipHcp (Arccos metric) — PRIORITY"),
        ("sand_hcp", "sandHcp (Arccos metric) — PRIORITY"),
    ]
    for col, label in metrics:
        vals = [v for v in (_num(r.get(col)) for r in rounds) if v is not None]
        avg = round(sum(vals) / len(vals), 2) if vals else None
        ws.append([label, avg if avg is not None else "n/a"])
        if col in short_game:
            for c in ws[ws.max_row]:
                c.fill, c.font = hl, bold

    # Career aggregate rates (Arccos tourAnalyticsSummary — robust vs per-round avg).
    ws.append([])
    ws.append(["CAREER AGGREGATE (Arccos)", "Value"])
    for c in ws[ws.max_row]:
        c.font, c.fill = bold, hdr
    kr = career.get("key_rates", {}) or {}
    for col, label, sgflag in [
        ("scramble_chip_save_pct", "Chip/short-game save % — PRIORITY", True),
        ("chip_down_pct", "Chip 'get it close' % — PRIORITY", True),
        ("chip_error_rate", "Chip error rate % — PRIORITY", True),
        ("sand_save_pct", "Sand save % — PRIORITY", True),
        ("gir_pct", "GIR % (career)", False),
        ("fairway_pct", "Fairway % (career)", False),
        ("three_putt_pct", "3-putt % (career)", False),
        ("one_putt_pct", "1-putt % (career)", False),
        ("putts_per_round", "Putts / round (career)", False),
    ]:
        v = kr.get(col)
        ws.append([label, v if v is not None else "n/a"])
        if sgflag:
            for c in ws[ws.max_row]:
                c.fill, c.font = hl, bold

    ws.append([])
    ws.append(["PROVENANCE & GLOSSARY"])
    ws[f"A{ws.max_row}"].font = bold
    for line in (
        "Source: UNOFFICIAL reverse-engineered Arccos API (api.arccosgolf.com), pulled by",
        "pull_arccos.py (custom stdlib client). Not affiliated with Arccos Golf LLC.",
        f"Date pulled: {pulled_at}  |  Rounds: {counts['rounds']}  Shots: {counts['shots']}",
        "",
        "REAL (native API): score, putts, GIR, fairway flags, penalties, clubs/terrain",
        "distances + dispersion + GIR%, Arccos category handicaps (drive/approach/chip/",
        "sand/putt Hcp — proprietary NEGATIVE scale, NOT USGA; clamped at -30).",
        "",
        "DERIVED (computed here, labeled _approx / _calibrated / _inferred):",
        "- par_calibrated: GPS hole length -> par class, nudged so holes sum to course par.",
        "- Strokes Gained (sg_*_approx): RECONSTRUCTED from shot start/end GPS distance-to-",
        f"  pin vs a published baseline ({BASELINE_SOURCE}). The official /v2/sga endpoint is",
        "  permission-gated (401) for dashboard tokens, so SG is rebuilt, NOT from Arccos.",
        "  Lie per shot is approximate (only tee/green are reliable) — treat SG as indicative.",
        "- scramble_pct: missed GIR + holed out <= par_calibrated. Native isUpDown flags were",
        "  unpopulated at pull time (kept verbatim as *_native).",
        "- approach_proximity_yd: end distance-to-pin of the last non-putt shot.",
        "Shot GPS coordinates ARE included (shots.csv/holes.csv/maps) per owner request.",
        "Identity (name, email, DOB, GHIN#) is still excluded from all outputs.",
    ):
        ws.append([line])
    wb.save(os.path.join(OUT_DIR, "arccos_sg_baseline.xlsx"))


# ---------------------------------------------------------------------------
# Discover (kept)
# ---------------------------------------------------------------------------

def _keys(o):
    if isinstance(o, dict):
        return sorted(o.keys())
    if isinstance(o, list) and o:
        return f"list[{len(o)}] -> {_keys(o[0])}"
    return type(o).__name__


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--fetch", action="store_true", help="Fetch endpoints into cache.")
    p.add_argument("--build", action="store_true", help="Build outputs from cache.")
    p.add_argument("--n", type=int, default=50, help="Max recent rounds (default 50).")
    p.add_argument("--after", type=str, default=None, help="Only rounds on/after YYYY-MM-DD.")
    p.add_argument("--login", action="store_true",
                   help="One-time email/password login -> stores accessKey (no DevTools).")
    p.add_argument("--include-gps", action="store_true",
                   help="include lat/lng columns in shots.csv/holes.csv (privacy-sensitive)")
    # Optional path overrides (defaults unchanged when flags absent). The creds file
    # is JSON: {"access_key": ..., "user_id": ...} (or bearer_token+user_id). When --out
    # is set it replaces the cache + discovery locations too (derived from OUT_DIR).
    p.add_argument("--creds", type=str, default=None,
                   help="Path to Arccos creds JSON (default: ~/.arccos_creds.json).")
    p.add_argument("--out", type=str, default=None,
                   help="Output directory (default: ./arccos_out, or this script's dir).")
    args = p.parse_args()

    global INCLUDE_GPS, CREDS_PATH, OUT_DIR, CACHE, DISCOVERY_DIR
    if args.include_gps:
        INCLUDE_GPS = True
    if args.creds:
        CREDS_PATH = os.path.expanduser(args.creds)
    if args.out:
        OUT_DIR = os.path.abspath(args.out)
        CACHE = os.path.join(OUT_DIR, "_cache_raw")
        DISCOVERY_DIR = os.path.join(OUT_DIR, "_discovery")

    if args.login:
        interactive_login()
        return

    pulled_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    do_fetch = args.fetch or not args.build
    do_build = args.build or not args.fetch

    if do_fetch:
        token, uid = load_creds()
        fetch_all(token, uid, args.n, args.after)
    if do_build:
        counts = build(pulled_at)
        print("\nBUILT:", ", ".join(f"{k}={v}" for k, v in counts.items()))
        print(f"Outputs -> {OUT_DIR}")


if __name__ == "__main__":
    main()
